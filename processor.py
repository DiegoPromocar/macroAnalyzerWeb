import openpyxl
import os
from datetime import datetime

# Directorio para guardar los reportes generados temporalmente
OUTPUT_DIR = 'generated_reports'
os.makedirs(OUTPUT_DIR, exist_ok=True)

def get_col_index(ws, col_name, row_index=4):
    for cell in ws[row_index]:
        if cell.value == col_name:
            return cell.column
    return None

def translate_months(ws, language, start_row=3):
    if language.lower() != 'english':
        return

    month_map = {
        "1": "January", "2": "February", "3": "March", "4": "April",
        "5": "May", "6": "June", "7": "July", "8": "August",
        "9": "September", "10": "October", "11": "November", "12": "December"
    }
    for row in ws.iter_rows(min_row=start_row, max_row=start_row):
        for cell in row:
            if cell.value and str(cell.value) in month_map:
                cell.value = month_map[str(cell.value)]

def find_last_month_info(ws, total_row_identifier):
    last_row = ws.max_row
    regist_cols = []
    for cell in ws[4]: # Fila 4 para buscar "Regist."
        if cell.value == "Regist.":
            regist_cols.append(cell.column)

    total_row_index = 0
    if isinstance(total_row_identifier, int):
        total_row_index = total_row_identifier
    else: # string identifier
        for i in range(1, last_row + 1):
            if ws.cell(row=i, column=1).value == total_row_identifier:
                total_row_index = i
                break
    
    if not total_row_index:
        return None, None

    matric_values = []
    for col_idx in regist_cols:
        cell_value = ws.cell(row=total_row_index, column=col_idx).value
        if cell_value is not None:
            matric_values.append({'value': cell_value, 'col': col_idx})

    # Lógica para encontrar el mes basado en matriculaciones duplicadas
    last_month_col = None
    if len(matric_values) > 1:
        value_counts = {}
        for item in matric_values:
            val = item['value']
            if val in value_counts:
                value_counts[val].append(item['col'])
            else:
                value_counts[val] = [item['col']]
        
        for val, cols in value_counts.items():
            if len(cols) > 1:
                last_month_col = cols[0] # El primer duplicado encontrado
                break
        
        # Si no hay duplicados, tomar el penúltimo
        if not last_month_col and len(matric_values) > 1:
            last_month_col = matric_values[-2]['col']

    elif len(matric_values) == 1:
        last_month_col = matric_values[0]['col']

    if last_month_col:
        return ws.cell(row=3, column=last_month_col).value, last_month_col
    
    return None, None


def generate_report(pais, anyo, mes, dia, idioma, segmentos, segmentation, modelos_path, marcas_path):
    """
    Procesa los archivos de Excel para generar un reporte consolidado.
    """
    # Cargar libros de trabajo
    wb_modelos = openpyxl.load_workbook(modelos_path)
    wb_marcas = openpyxl.load_workbook(marcas_path)

    # --- 1. Procesar Libro de Marcas ---
    ws_marcas = wb_marcas.active
    translate_months(ws_marcas, idioma, start_row=3)
    
    # Encontrar la fila "Total" para la búsqueda del último mes
    n_fil_marcas = 0
    for i in range(1, ws_marcas.max_row + 1):
        if str(ws_marcas.cell(row=i, column=1).value).strip() == "Total":
            n_fil_marcas = i
            break
            
    if n_fil_marcas == 0: # Si no se encuentra "Total", usar la antepenúltima fila como fallback
        n_fil_marcas = ws_marcas.max_row - 2 if ws_marcas.max_row > 2 else 1

    ult_mes_marcas, _ = find_last_month_info(ws_marcas, n_fil_marcas)


    # --- 2. Procesar Libro de Modelos ---
    sheets_to_process = [sheet for sheet in wb_modelos.sheetnames if sheet.lower() != 'general']
    
    for sheet_name in sheets_to_process:
        ws = wb_modelos[sheet_name]
        translate_months(ws, idioma, start_row=3)
        find_last_month_info(ws, "Total General")

    # Eliminar la hoja "General" si existe
    if 'General' in wb_modelos.sheetnames:
        del wb_modelos['General']
        
    # --- Lógica de la plantilla y copia de hojas (a implementar) ---
    # Por ahora, solo guardamos el libro de modelos modificado.
    
    # --- 4. Guardar el resultado ---
    segment_code = "PC" if "Passenger Cars" in segmentos else "LCV"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Construir el nombre del archivo de salida
    if segmentation == "VOLKSWAGEN":
        output_filename = f"BENEF_{pais}_{anyo}-{int(mes):02d}_{segment_code}_VW_E_{timestamp}.xlsx"
    else:
        output_filename = f"BENEF_{pais}_{anyo}-{int(mes):02d}_{segment_code}_E_{timestamp}.xlsx"
        
    output_path = os.path.join(OUTPUT_DIR, output_filename)

    wb_modelos.save(output_path)

    return output_path